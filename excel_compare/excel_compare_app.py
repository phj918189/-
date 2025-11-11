import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tksheet import Sheet, rounded_box_coords
import pandas as pd
import openpyxl
import xlrd
from openpyxl import Workbook
import os
import json
from decimal import Decimal, InvalidOperation
from pathlib import Path
import threading


# ============================================
# 🔹 오래된 .xls → .xlsx 자동 변환 함수
# ============================================
def convert_xls_to_xlsx(xls_path):
    """
    오래된 .xls 파일을 openpyxl에서 읽을 수 있는 .xlsx로 변환
    
    Args:
        xls_path: 변환할 .xls 파일 경로
        
    Returns:
        변환된 .xlsx 파일 경로
        
    Raises:
        FileNotFoundError: 파일을 찾을 수 없는 경우
        xlrd.XLRDError: xlrd로 파일을 읽을 수 없는 경우
        PermissionError: 파일 쓰기 권한이 없는 경우
    """
    if not os.path.exists(xls_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {xls_path}")
    
    try:
        wb_xls = xlrd.open_workbook(xls_path)
    except Exception as e:
        raise xlrd.XLRDError(f"xls 파일 읽기 실패: {e}")
    
    try:
        wb_xlsx = Workbook()
        # 모든 시트 변환 (첫 번째 시트만이 아닌)
        for sheet_idx, sheet_xls in enumerate(wb_xls.sheets()):
            if sheet_idx == 0:
                ws_xlsx = wb_xlsx.active
                ws_xlsx.title = sheet_xls.name
            else:
                ws_xlsx = wb_xlsx.create_sheet(title=sheet_xls.name)
            
            # 셀 데이터 복사
            for r in range(sheet_xls.nrows):
                for c in range(sheet_xls.ncols):
                    cell_value = sheet_xls.cell_value(r, c)
                    ws_xlsx.cell(row=r + 1, column=c + 1).value = cell_value
        
        new_path = os.path.splitext(xls_path)[0] + "_converted.xlsx"
        wb_xlsx.save(new_path)
        return new_path
    except PermissionError:
        raise PermissionError(f"파일 쓰기 권한이 없습니다: {new_path}")
    except Exception as e:
        raise Exception(f"xlsx 파일 변환 중 오류 발생: {e}")


# ============================================
# 🔹 설정 관리 클래스
# ============================================
class ConfigManager:
    """설정 파일 관리 클래스"""
    CONFIG_FILE = "excel_compare_config.json"
    
    # 기본 설정
    DEFAULT_CONFIG = {
        "window": {
            "width": 1700,
            "height": 850,
            "x": None,  # None이면 화면 중앙
            "y": None
        },
        "theme": "light",  # "light" 또는 "dark"
        "themes": {
            "light": {
                "diff_highlight": "green",
                "button_primary": "#3C91E6",
                "button_secondary": "#FFC107",
                "label_bg": "#F0F0F0",
                "text_secondary": "#666666",
                "button_text": "white",
                "button_secondary_text": "black",
                "bg": "white",
                "fg": "black"
            },
            "dark": {
                "diff_highlight": "#FF6B6B",
                "button_primary": "#4A90E2",
                "button_secondary": "#F5A623",
                "label_bg": "#2C2C2C",
                "text_secondary": "#CCCCCC",
                "button_text": "white",
                "button_secondary_text": "black",
                "bg": "#1E1E1E",
                "fg": "#E0E0E0"
            }
        },
        "font": {
            "family": "맑은 고딕",
            "size_default": 10,
            "size_path": 9,
            "weight_default": "bold"
        }
    }
    
    @classmethod
    def load_config(cls):
        """설정 파일 로드"""
        config_path = Path(cls.CONFIG_FILE)
        if config_path.exists():
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    user_config = json.load(f)
                # 기본 설정과 병합
                config = cls.DEFAULT_CONFIG.copy()
                cls._deep_update(config, user_config)
                return config
            except Exception as e:
                print(f"설정 파일 로드 실패, 기본 설정 사용: {e}")
                return cls.DEFAULT_CONFIG.copy()
        else:
            # 기본 설정으로 파일 생성
            cls.save_config(cls.DEFAULT_CONFIG)
            return cls.DEFAULT_CONFIG.copy()
    
    @classmethod
    def save_config(cls, config):
        """설정 파일 저장"""
        try:
            config_path = Path(cls.CONFIG_FILE)
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"설정 파일 저장 실패: {e}")
    
    @staticmethod
    def _deep_update(base_dict, update_dict):
        """딕셔너리 깊은 병합"""
        for key, value in update_dict.items():
            if key in base_dict and isinstance(base_dict[key], dict) and isinstance(value, dict):
                ConfigManager._deep_update(base_dict[key], value)
            else:
                base_dict[key] = value


# ============================================
# 🔹 엑셀 비교 GUI
# ============================================
class ExcelComparatorApp:
    
    def __init__(self, root):
        self.root = root
        self.root.title("엑셀 비교 뷰어")
        
        # 설정 로드
        self.config = ConfigManager.load_config()
        self.theme = self.config["theme"]
        self.theme_colors = self.config["themes"][self.theme]
        
        # 창 크기/위치 설정
        self._setup_window_geometry()
        
        # 창 닫기 이벤트 바인딩 (설정 저장)
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)

        self.df_left = None
        self.df_right = None
        self.path_left = None
        self.path_right = None
        self._syncing_scroll = False
        self._syncing_select = False
        self._compare_cancelled = False  # 비교 취소 플래그
        self._diff_list = []  # 차이점 목록 저장
        
        # 테마 적용
        self._apply_theme()

        # ─ 상단 버튼 ─
        top = tk.Frame(root, bg=self.theme_colors["bg"])
        top.pack(pady=10)
        self.top_frame = top

        tk.Button(top, text="📂 첫 번째 파일", command=self.load_left,
                  bg=self.theme_colors["button_primary"], 
                  fg=self.theme_colors["button_text"], width=15).grid(row=0, column=0, padx=8)
        tk.Button(top, text="📂 두 번째 파일", command=self.load_right,
                  bg=self.theme_colors["button_primary"], 
                  fg=self.theme_colors["button_text"], width=15).grid(row=0, column=1, padx=8)
        # 비교 버튼 프레임 (비교 시작 + 취소)
        compare_frame = tk.Frame(top)
        compare_frame.grid(row=0, column=2, padx=8)
        
        self.compare_button = tk.Button(compare_frame, text="🔍 비교 시작", command=self.compare_files,
                  bg=self.theme_colors["button_secondary"], 
                  fg=self.theme_colors["button_secondary_text"], width=10)
        self.compare_button.pack()
        
        self.cancel_button = tk.Button(compare_frame, text="❌ 비교 취소", command=self._cancel_compare,
                  bg="#DC3545", fg="white", width=15, state=tk.DISABLED)
        # self.cancel_button.pack(pady=(5, 0))
        
        # 테마 전환 버튼
        theme_text = "🌙 다크모드" if self.theme == "light" else "☀️ 라이트모드"
        tk.Button(top, text=theme_text, command=self._toggle_theme,
                  bg=self.theme_colors["label_bg"], 
                  fg=self.theme_colors["fg"], width=15).grid(row=0, column=4, padx=8)
        
        # ─ 셀 주소 표시 라벨 ─
        font_config = self.config["font"]
        font_default = (font_config["family"], font_config["size_default"], font_config["weight_default"])
        self.cell_address_label = tk.Label(top, text="셀 주소: ", 
                                           font=font_default, 
                                           bg=self.theme_colors["label_bg"],
                                           fg=self.theme_colors["fg"],
                                           padx=10, pady=5)
        self.cell_address_label.grid(row=0, column=3, padx=8)

        # ─ 시트 영역 ─
        frame = tk.Frame(root, bg=self.theme_colors["bg"])
        frame.pack(padx=10, pady=10, fill="both", expand=True)

        self.left_sheet = Sheet(frame, width=830, height=720)
        self.right_sheet = Sheet(frame, width=830, height=720)
        self.left_sheet.grid(row=0, column=0, sticky="nsew", padx=5)
        self.right_sheet.grid(row=0, column=1, sticky="nsew", padx=5)

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_columnconfigure(1, weight=1)

        # ─ 동기화 설정 ─
        self._install_sync_scroll()
        self._install_sync_selection()

        # ─ 시트 선택 드롭다운 및 시트 데이터 맵 ─
        self.left_sheet_selector = None
        self.right_sheet_selector = None
        self.common_sheet_selector = None

        self.left_sheets_map = {}
        self.right_sheets_map = {}
        self.left_sheet_name = None
        self.right_sheet_name = None

    # ============================================
    # 🔹 엑셀 데이터 B2부터 불러오기 (A열은 빈 열로 유지)
    # ============================================
    def _load_b2_from_excel(self, path):
        """
        A열과 1행은 빈 셀, B2부터 데이터 시작
        A열도 포함해서 읽되, A열은 빈 값으로 유지
        1행도 빈 행으로 추가
        """
        wb = openpyxl.load_workbook(path, data_only=True)

        # 항상 첫 시트가 아닌 "채취일지" 시트를 사용
        ws = wb["채취일지"] if "채취일지" in wb.sheetnames else wb.active
        return self._load_b2_from_worksheet(ws)

    def _load_b2_from_worksheet(self, ws):
        """
        주어진 worksheet에서 B2 기준으로 데이터를 DataFrame으로 반환
        """
        # 병합된 셀 정보 수집
        # 병합된 셀의 주 셀(첫 번째 셀)에만 값을 저장하고, 나머지 병합된 위치는 빈 값으로 처리
        merged_cell_map = {}  # Excel (행, 열) -> 값 (주 셀만 저장)
        merged_cell_ranges = {}  # Excel (행, 열) -> 병합 범위 정보 (주 셀만 저장)
        merged_cell_ignore = set()  # 병합된 셀 범위 내의 하위 셀들 (무시해야 할 위치)
        
        for merged_range in ws.merged_cells.ranges:
            # 병합 범위: min_row, min_col, max_row, max_col
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            
            # 주 셀(첫 번째 셀)의 값 가져오기
            master_cell = ws.cell(min_row, min_col)
            master_value = master_cell.value if master_cell.value is not None else ""
            
            # 주 셀에만 값 저장 (B2부터만 처리)
            if min_row >= 2 and min_col >= 2:  # B2부터만 처리
                merged_cell_map[(min_row, min_col)] = str(master_value)
                merged_cell_ranges[(min_row, min_col)] = (min_row, min_col, max_row, max_col)
            
            # 병합된 셀 범위 내의 하위 셀들(주 셀 제외)은 무시 목록에 추가
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if r >= 2 and c >= 2:  # B2부터만 처리
                        # 주 셀이 아닌 경우에만 무시 목록에 추가
                        if not (r == min_row and c == min_col):
                            merged_cell_ignore.add((r, c))
        
        # Excel의 실제 행/열 범위 확인 (원본 구조 유지를 위해)
        excel_max_row = ws.max_row  # Excel의 최대 행 번호
        excel_max_col = ws.max_column  # Excel의 최대 열 번호
        
        data = []
        # Excel의 모든 행을 순회 (2행부터, 빈 행도 포함)
        for excel_row in range(2, excel_max_row + 1):
            values = []
            # A열(첫 번째 열)은 항상 빈 문자열로 처리
            values.append("")
            
            # B열부터 Excel의 최대 열까지 순회 (빈 열도 포함)
            for excel_col in range(2, excel_max_col + 1):  # 2=B열부터
                # 병합된 셀 범위 내의 하위 셀인지 확인 (무시해야 할 위치)
                if (excel_row, excel_col) in merged_cell_ignore:
                    # 병합된 셀의 하위 셀은 빈 값으로 처리 (원본 구조 유지)
                    values.append("")
                elif (excel_row, excel_col) in merged_cell_map:
                    # 병합된 셀의 주 셀인 경우, 저장된 값 사용
                    values.append(merged_cell_map[(excel_row, excel_col)])
                else:
                    # 일반 셀 값 (셀이 없거나 None이면 빈 문자열)
                    cell = ws.cell(excel_row, excel_col)
                    values.append(str(cell.value) if cell.value is not None else "")
            
            # 모든 행 추가 (빈 행도 포함하여 원본 구조 유지)
            data.append(values)
        
        # 열 개수 계산 (A열 + B열부터 최대 열까지 = excel_max_col)
        max_cols = excel_max_col  # A열(1개) + B열부터 최대 열까지(excel_max_col-1개) = excel_max_col개
        
        # DataFrame 생성 (빈 행/열도 모두 포함하여 원본 구조 유지)
        if data:
            df = pd.DataFrame(data)
            # 모든 열 유지 (빈 열도 포함하여 원본 구조 유지)
            # 열 개수가 부족하면 빈 열 추가
            if df.shape[1] < max_cols:
                # 부족한 열만큼 빈 열 추가
                for _ in range(max_cols - df.shape[1]):
                    df[len(df.columns)] = ""
        else:
            # 데이터가 없어도 Excel의 열 범위는 유지
            df = pd.DataFrame([[""] * max_cols])
        
        # 1행 추가 (모두 빈 값, A열 포함)
        empty_first_row = [""] * max_cols
        df = pd.concat([pd.DataFrame([empty_first_row], columns=df.columns), df], ignore_index=True)
        
        return df.reset_index(drop=True)

    def _load_all_sheets_from_excel(self, path):
        """
        모든 시트를 읽어 dict[시트명] = DataFrame 형태로 반환
        """
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets_map = {}
        for ws in wb.worksheets:
            try:
                sheets_map[ws.title] = self._load_b2_from_worksheet(ws)
            except Exception as e:
                print(f"시트 '{ws.title}' 로드 실패: {e}")
        active_name = wb.active.title if wb.active else (wb.sheetnames[0] if wb.sheetnames else None)
        return sheets_map, active_name

    # ============================================
    # 🔹 스크롤 동기화 데코레이터 팩토리
    # ============================================
    def _create_scroll_sync_decorator(self, target_moveto_func):
        """
        스크롤 동기화 데코레이터 생성 (공통 로직 추출)
        
        Args:
            target_moveto_func: 대상 시트의 moveto 함수
            
        Returns:
            데코레이터 함수
        """
        def decorator(source_view_func):
            """원본 뷰 함수를 래핑하는 데코레이터"""
            def wrapper(*args, **kwargs):
                # 원본 함수 실행 (스크롤 이동)
                source_view_func(*args, **kwargs)
                
                # 동기화 중이 아니면 대상 시트도 동일한 위치로 이동
                if not self._syncing_scroll:
                    try:
                        # 현재 스크롤 위치 가져오기
                        current_position, _ = source_view_func()
                        
                        # 대상 시트 동기화
                        self._syncing_scroll = True
                        target_moveto_func("moveto", current_position)
                        self._syncing_scroll = False
                    except AttributeError as e:
                        # 뷰 함수가 호출 가능하지 않거나 속성이 없는 경우
                        # tksheet의 내부 구조 변경 시 발생할 수 있음
                        pass
                    except TypeError as e:
                        # 잘못된 인자 타입 (예: None 전달)
                        pass
                    except ValueError as e:
                        # 잘못된 값 (예: 범위를 벗어난 위치)
                        pass
                    except (IndexError, KeyError) as e:
                        # 튜플 언패킹 실패 등
                        pass
                    finally:
                        # 동기화 플래그는 항상 해제
                        self._syncing_scroll = False
            
            return wrapper
        return decorator

    # ============================================
    # 🔹 스크롤 동기화 설치
    # ============================================
    def _install_sync_scroll(self):
        """
        양쪽 시트의 스크롤을 동기화
        데코레이터 패턴을 사용하여 깔끔하게 구현
        """
        # 원본 뷰 함수 백업 (복원 가능하도록)
        self._left_yview_orig = self.left_sheet.MT.yview
        self._right_yview_orig = self.right_sheet.MT.yview
        self._left_xview_orig = self.left_sheet.MT.xview
        self._right_xview_orig = self.right_sheet.MT.xview

        # 데코레이터 생성
        sync_left_to_right_y = self._create_scroll_sync_decorator(
            self.right_sheet.MT.yview_moveto
        )
        sync_right_to_left_y = self._create_scroll_sync_decorator(
            self.left_sheet.MT.yview_moveto
        )
        sync_left_to_right_x = self._create_scroll_sync_decorator(
            self.right_sheet.MT.xview_moveto
        )
        sync_right_to_left_x = self._create_scroll_sync_decorator(
            self.left_sheet.MT.xview_moveto
        )

        # 데코레이터 적용
        self.left_sheet.MT.yview = sync_left_to_right_y(self._left_yview_orig)
        self.right_sheet.MT.yview = sync_right_to_left_y(self._right_yview_orig)
        self.left_sheet.MT.xview = sync_left_to_right_x(self._left_xview_orig)
        self.right_sheet.MT.xview = sync_right_to_left_x(self._right_xview_orig)

    # ============================================
    # 🔹 Excel 열/주소 변환 유틸리티 함수들
    # ============================================
    @staticmethod
    def _excel_col_num_to_name(col_num):
        """
        Excel 열 번호를 열 이름으로 변환 (1=A, 2=B, ..., 26=Z, 27=AA, ...)
        
        공통 유틸리티 함수로, 여러 곳에서 사용됨:
        - _row_col_to_excel_address: 셀 주소 변환
        - _generate_headers: 열 헤더 생성
        
        Args:
            col_num: Excel 열 번호 (1-based)
            
        Returns:
            Excel 열 이름 (예: "A", "B", "AA", "ZZ")
        """
        num = col_num
        result = ""
        while num > 0:
            num, remainder = divmod(num - 1, 26)
            result = chr(65 + remainder) + result
        return result
    
    def _row_col_to_excel_address(self, row, col):
        """
        tksheet의 row/col 인덱스(0-based)를 Excel 주소로 변환
        - row=0 → Excel 행 1 (빈 행)
        - row=1 → Excel 행 2 (데이터 시작)
        - col=0 → Excel A열
        - col=1 → Excel B열
        
        Args:
            row: tksheet 행 인덱스 (0-based)
            col: tksheet 열 인덱스 (0-based)
            
        Returns:
            Excel 주소 문자열 (예: "A1", "B2", "AA34", "ZZ100")
        """
        # Excel 행 번호 (1-based): 화면 row + 1
        excel_row = row + 1
        
        # Excel 열 이름 변환 (공통 함수 사용)
        excel_col_num = col + 1  # Excel 열 번호 (1=A, 2=B, ...)
        col_name = self._excel_col_num_to_name(excel_col_num)
        
        return f"{col_name}{excel_row}"
    
    @staticmethod
    def _generate_headers(num_columns, excel_start_col=1):
        """
        Excel 열 헤더 생성 (예: A, B, C, D, ...)
        
        A열은 빈 열이지만 화면에는 표시되어야 함
        - DataFrame[0] = Excel A열 → 헤더 "A" (빈 열)
        - DataFrame[1] = Excel B열 → 헤더 "B" (데이터 시작)
        
        Args:
            num_columns: 생성할 헤더 개수 (DataFrame의 열 개수)
            excel_start_col: Excel 열 번호 (1=A, 2=B, 3=C, ...)
                            A열부터 표시하므로 기본값은 1
        """
        headers = []
        # Excel 열 번호를 Excel 열 이름으로 변환 (공통 함수 사용)
        for excel_col_num in range(excel_start_col, excel_start_col + num_columns):
            headers.append(ExcelComparatorApp._excel_col_num_to_name(excel_col_num))
        return headers

    # ============================================
    # 🔹 셀 선택 동기화 + 주소 표시
    # ============================================
    def _install_sync_selection(self):
        def sync_from_to(source, target):
            if self._syncing_select:
                return
            selected = source.get_currently_selected()
            if not selected or len(selected) < 2:
                return
            r, c = selected[0], selected[1]
            self._syncing_select = True
            try:
                target.select_cell(r, c, redraw=True)
                target.see(r, c)
            finally:
                self._syncing_select = False

        def update_cell_address(sheet, side):
            """선택된 셀의 Excel 주소를 업데이트"""
            try:
                selected = sheet.get_currently_selected()
                if selected and len(selected) >= 2:
                    r, c = selected[0], selected[1]
                    address = self._row_col_to_excel_address(r, c)
                    self.cell_address_label.config(text=f"셀 주소: {address} ({side})")
                else:
                    self.cell_address_label.config(text="셀 주소: -")
            except Exception:
                pass

        def on_left_select(event=None):
            sync_from_to(self.left_sheet, self.right_sheet)
            update_cell_address(self.left_sheet, "왼쪽")

        def on_right_select(event=None):
            sync_from_to(self.right_sheet, self.left_sheet)
            update_cell_address(self.right_sheet, "오른쪽")

        for ev in ("<ButtonRelease-1>", "<KeyRelease>", "<B1-Motion>"):
            self.left_sheet.MT.bind(ev, on_left_select, add=True)
            self.right_sheet.MT.bind(ev, on_right_select, add=True)

        # 람다 클로저 이슈 방지를 위해 명시적 함수 사용
        def create_left_select_handler():
            """왼쪽 시트 선택 핸들러 생성"""
            def handler(p):
                on_left_select()
            return handler
        
        def create_right_select_handler():
            """오른쪽 시트 선택 핸들러 생성"""
            def handler(p):
                on_right_select()
            return handler
        
        for ev in ("cell_select", "row_select", "column_select", "shift_cell_select", "drag_select"):
            self.left_sheet.extra_bindings(ev, create_left_select_handler())
            self.right_sheet.extra_bindings(ev, create_right_select_handler())

    # ============================================
    # 🔹 진행 표시 다이얼로그 생성/업데이트
    # ============================================
    def _show_progress_dialog(self, message):
        """진행 표시 다이얼로그 생성"""
        dialog = tk.Toplevel(self.root)
        dialog.title("파일 로딩 중...")
        dialog.geometry("400x120")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 중앙 배치
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (dialog.winfo_screenheight() // 2) - (120 // 2)
        dialog.geometry(f"400x120+{x}+{y}")
        
        # 진행 메시지
        label = tk.Label(dialog, text=message, font=("맑은 고딕", 10), pady=10)
        label.pack()
        
        # 진행 바
        progress = ttk.Progressbar(dialog, mode='indeterminate', length=300)
        progress.pack(pady=10)
        progress.start(10)  # 애니메이션 시작
        
        dialog.update()
        return dialog, progress, label
    
    def _update_progress_message(self, dialog, label, message):
        """진행 메시지 업데이트"""
        if dialog.winfo_exists():
            label.config(text=message)
            dialog.update()
    
    def _close_progress_dialog(self, dialog, progress):
        """진행 표시 다이얼로그 닫기"""
        if dialog.winfo_exists():
            progress.stop()
            dialog.destroy()

    # ============================================
    # 🔹 파일 불러오기 (.xls 자동 변환 + B2 기준) - 비동기 버전
    # ============================================
    def _load_file_common(self, side):
        """
        파일 로드 공통 로직 (비동기 처리)
        
        Args:
            side: "left" 또는 "right"
        """
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")])
        if not path:
            return False
        
        # 파일 경로 표시 업데이트 (로딩 시작)
        filename = os.path.basename(path)
        self._update_file_path_label(side, path, loading=True)
        
        # 진행 다이얼로그 먼저 생성 (메인 스레드에서)
        dialog, progress, label = self._show_progress_dialog(
            f"{'왼쪽' if side == 'left' else '오른쪽'} 파일 로딩 중..."
        )
        
        # 비동기 로딩 시작
        thread = threading.Thread(
            target=self._load_file_async,
            args=(side, path, dialog, progress, label),
            daemon=True
        )
        thread.start()
        
        return True
    
    def _load_file_async(self, side, path, dialog, progress, label):
        """
        파일 로드 비동기 처리
        
        Args:
            side: "left" 또는 "right"
            path: 파일 경로
            dialog: 진행 다이얼로그
            progress: 진행 바
            label: 진행 메시지 라벨
        """
        try:
            # .xls 파일 변환
            if path.lower().endswith(".xls"):
                self.root.after(0, lambda: self._update_progress_message(
                    dialog, label, "파일 변환 중... (.xls → .xlsx)"
                ))
                try:
                    path = convert_xls_to_xlsx(path)
                except Exception as e:
                    self.root.after(0, lambda: messagebox.showerror(
                        "오류", f"파일 변환 실패:\n{e}"
                    ))
                    self.root.after(0, lambda: self._close_progress_dialog(dialog, progress))
                    return
            
            # 파일 로드
            self.root.after(0, lambda: self._update_progress_message(
                dialog, label, "엑셀 파일 읽는 중... (모든 시트)"
            ))
            sheets_map, active_name = self._load_all_sheets_from_excel(path)

            def resolve_default_sheet():
                preferred = self.config.get("preferred_sheet") if isinstance(self.config, dict) else None
                if preferred and preferred in sheets_map:
                    return preferred
                if "채취일지" in sheets_map:
                    return "채취일지"
                if active_name and active_name in sheets_map:
                    return active_name
                return next(iter(sheets_map), None)

            default_sheet = resolve_default_sheet()

            # 시트 맵 저장 및 UI 업데이트 (메인 스레드에서 실행)
            self.root.after(0, self._apply_loaded_sheets, side, path, sheets_map, default_sheet, dialog, progress)
            
        except Exception as e:
            # 오류 발생 시 메인 스레드에서 처리
            error_msg = f"{'왼쪽' if side == 'left' else '오른쪽'} 파일 로드 실패:\n{e}"
            self.root.after(0, lambda: messagebox.showerror("오류", error_msg))
            self.root.after(0, lambda: self._close_progress_dialog(dialog, progress))
    
    def _set_all_cells_center_alignment(self, sheet, num_rows, num_cols):
        """
        모든 셀에 가운데 정렬 설정
        
        Args:
            sheet: tksheet Sheet 객체
            num_rows: 행 개수
            num_cols: 열 개수
        """
        try:
            # 방법 1: table_align 메서드로 전체 테이블 정렬 (가장 간단한 방법)
            if hasattr(sheet, 'table_align'):
                try:
                    sheet.table_align("center", redraw=True)
                    return
                except:
                    try:
                        sheet.table_align("center")
                        return
                    except Exception as e:
                        print(f"table_align 실패: {e}")
            
            # 방법 2: align_cells 메서드로 모든 셀 정렬
            if hasattr(sheet, 'align_cells'):
                try:
                    # 모든 셀의 좌표 리스트 생성
                    cells = [(row, col) for row in range(num_rows) for col in range(num_cols)]
                    sheet.align_cells(cells, align="center", redraw=True)
                    return
                except:
                    try:
                        cells = [(row, col) for row in range(num_rows) for col in range(num_cols)]
                        sheet.align_cells(cells, align="center")
                        return
                    except Exception as e:
                        print(f"align_cells 실패: {e}")
            
            # 방법 3: align_columns 메서드로 모든 열 정렬
            if hasattr(sheet, 'align_columns'):
                try:
                    columns = list(range(num_cols))
                    sheet.align_columns(columns, align="center", redraw=True)
                    return
                except:
                    try:
                        columns = list(range(num_cols))
                        sheet.align_columns(columns, align="center")
                        return
                    except Exception as e:
                        print(f"align_columns 실패: {e}")
            
            # 방법 4: align_rows 메서드로 모든 행 정렬
            if hasattr(sheet, 'align_rows'):
                try:
                    rows = list(range(num_rows))
                    sheet.align_rows(rows, align="center", redraw=True)
                    return
                except:
                    try:
                        rows = list(range(num_rows))
                        sheet.align_rows(rows, align="center")
                        return
                    except Exception as e:
                        print(f"align_rows 실패: {e}")
            
            # 방법 5: 개별 셀별로 align 메서드 사용
            if hasattr(sheet, 'align'):
                for row in range(num_rows):
                    for col in range(num_cols):
                        try:
                            sheet.align(row, col, align="center")
                        except:
                            try:
                                sheet.align(row, col, "center")
                            except:
                                pass
                # redraw 호출
                try:
                    sheet.refresh()
                except:
                    pass
                return
                
        except Exception as e:
            # 정렬 설정 실패해도 계속 진행 (오류 무시)
            print(f"셀 정렬 설정 실패 (무시): {e}")
    
    def _update_file_path_label(self, side, path, loading=False):
        """
        파일 경로 라벨 업데이트
        
        Args:
            side: "left" 또는 "right"
            path: 파일 경로
            loading: 로딩 중 여부
        """
        filename = os.path.basename(path)
        full_path = path  # 전체 경로도 표시
        font_config = self.config["font"]
        font_path = (font_config["family"], font_config["size_path"])
        
        # 로딩 중이면 상태 표시
        status_text = " (로딩 중...)" if loading else ""
        display_text = f"{'왼쪽' if side == 'left' else '오른쪽'} 파일: {filename}{status_text}"
        
        if side == "left":
            if not hasattr(self, 'left_path_label'):
                self.left_path_label = tk.Label(self.root, text="", font=font_path, 
                                               fg=self.theme_colors["text_secondary"], 
                                               bg=self.theme_colors["bg"],
                                               anchor="w")
                self.left_path_label.pack(side="top", fill="x", padx=10, pady=(0, 5))
            self.left_path_label.config(text=display_text)
            # 툴팁으로 전체 경로 표시
            self.left_path_label.config(cursor="hand2")
            self.left_path_label.bind("<Enter>", lambda e: self._show_path_tooltip(e, full_path))
            self.left_path_label.bind("<Leave>", lambda e: self._hide_tooltip())
        else:
            if not hasattr(self, 'right_path_label'):
                self.right_path_label = tk.Label(self.root, text="", font=font_path, 
                                                fg=self.theme_colors["text_secondary"],
                                                bg=self.theme_colors["bg"],
                                                anchor="w")
                self.right_path_label.pack(side="top", fill="x", padx=10, pady=(0, 5))
            self.right_path_label.config(text=display_text)
            # 툴팁으로 전체 경로 표시
            self.right_path_label.config(cursor="hand2")
            self.right_path_label.bind("<Enter>", lambda e: self._show_path_tooltip(e, full_path))
            self.right_path_label.bind("<Leave>", lambda e: self._hide_tooltip())
    
    # ============================================
    # 🔹 시트 선택 드롭다운 처리
    # ============================================
    def _ensure_sheet_selector(self, side):
        """필요 시 좌/우 시트 콤보박스 생성"""
        if side == "left":
            if self.left_sheet_selector is None:
                self.left_sheet_selector = ttk.Combobox(self.top_frame, state="readonly", width=15)
                self.left_sheet_selector.grid(row=1, column=0, padx=8, pady=(4, 0), sticky="w")
                self.left_sheet_selector.bind("<<ComboboxSelected>>", lambda e: self._on_sheet_select("left"))
        else:
            if self.right_sheet_selector is None:
                self.right_sheet_selector = ttk.Combobox(self.top_frame, state="readonly", width=15)
                self.right_sheet_selector.grid(row=1, column=1, padx=8, pady=(4, 0), sticky="w")
                self.right_sheet_selector.bind("<<ComboboxSelected>>", lambda e: self._on_sheet_select("right"))

    def _set_sheet_selector_items(self, side, items, selected):
        """콤보박스 값 세팅"""
        self._ensure_sheet_selector(side)
        target_selector = self.left_sheet_selector if side == "left" else self.right_sheet_selector
        target_selector["values"] = items
        if selected and selected in items:
            target_selector.set(selected)
        elif items:
            target_selector.set(items[0])
        else:
            target_selector.set("")

    def _display_dataframe_on_sheet(self, sheet_widget, df):
        """DataFrame을 지정 시트 위젯에 출력"""
        sheet_widget.set_sheet_data(df.astype(str).values.tolist())
        sheet_widget.set_all_cell_sizes_to_text()
        sheet_widget.headers(self._generate_headers(df.shape[1]))
        sheet_widget.enable_bindings(("single_select", "drag_select", "copy", "arrowkeys"))
        self._set_all_cells_center_alignment(sheet_widget, df.shape[0], df.shape[1])

    def _update_current_df_and_display(self, side, sheet_name):
        """선택 시트에 따라 df 및 화면 갱신"""
        if side == "left":
            if sheet_name in self.left_sheets_map:
                self.left_sheet_name = sheet_name
                self.df_left = self.left_sheets_map[sheet_name]
                self._display_dataframe_on_sheet(self.left_sheet, self.df_left)
        else:
            if sheet_name in self.right_sheets_map:
                self.right_sheet_name = sheet_name
                self.df_right = self.right_sheets_map[sheet_name]
                self._display_dataframe_on_sheet(self.right_sheet, self.df_right)
        self._sync_common_selector_selection()

    def _on_sheet_select(self, side):
        """좌/우 시트 콤보박스 선택 변경"""
        if side == "left" and self.left_sheet_selector is not None:
            sel = self.left_sheet_selector.get()
            self._update_current_df_and_display("left", sel)
        elif side == "right" and self.right_sheet_selector is not None:
            sel = self.right_sheet_selector.get()
            self._update_current_df_and_display("right", sel)

    def _ensure_common_sheet_selector(self):
        """공통 시트 콤보박스 준비"""
        if self.common_sheet_selector is None:
            self.common_sheet_selector = ttk.Combobox(self.top_frame, state="readonly", width=15)
            self.common_sheet_selector.grid(row=1, column=2, padx=8, pady=(4, 0))
            self.common_sheet_selector.bind("<<ComboboxSelected>>", self._on_common_sheet_select)

    def _update_common_sheet_selector(self):
        """좌/우 공통 시트 목록 갱신"""
        common = sorted(set(self.left_sheets_map.keys()) & set(self.right_sheets_map.keys()))
        if not common:
            if self.common_sheet_selector is not None:
                self.common_sheet_selector["values"] = []
                self.common_sheet_selector.set("")
                self.common_sheet_selector.state(["disabled"])
            return

        self._ensure_common_sheet_selector()
        self.common_sheet_selector.state(["!disabled"])
        self.common_sheet_selector["values"] = common

        preferred = None
        if self.left_sheet_name == self.right_sheet_name and self.left_sheet_name in common:
            preferred = self.left_sheet_name
        else:
            current = self.common_sheet_selector.get()
            if current in common:
                preferred = current
            else:
                preferred = common[0]
        if preferred:
            self.common_sheet_selector.set(preferred)

    def _sync_common_selector_selection(self):
        """현재 좌/우 선택 상태에 맞춰 공통 콤보박스 값 유지"""
        if self.common_sheet_selector is None:
            return
        if self.left_sheet_name == self.right_sheet_name and self.left_sheet_name is not None:
            if self.left_sheet_name in set(self.common_sheet_selector["values"]):
                self.common_sheet_selector.set(self.left_sheet_name)
        else:
            current = self.common_sheet_selector.get()
            if current not in set(self.common_sheet_selector["values"]):
                self.common_sheet_selector.set("")

    def _on_common_sheet_select(self, event=None):
        """공통 시트 콤보박스 선택"""
        if self.common_sheet_selector is None:
            return
        selected = self.common_sheet_selector.get()
        if not selected:
            return
        if self.left_sheet_selector and selected in self.left_sheet_selector["values"]:
            self.left_sheet_selector.set(selected)
            self._update_current_df_and_display("left", selected)
        if self.right_sheet_selector and selected in self.right_sheet_selector["values"]:
            self.right_sheet_selector.set(selected)
            self._update_current_df_and_display("right", selected)

    def _apply_loaded_sheets(self, side, path, sheets_map, default_sheet_name, dialog, progress):
        """모든 시트 로드 후 UI에 반영"""
        try:
            sheet_names = list(sheets_map.keys())

            if side == "left":
                self.path_left = path
                self.left_sheets_map = sheets_map
                self._set_sheet_selector_items("left", sheet_names, default_sheet_name)
                if sheet_names:
                    target = default_sheet_name if default_sheet_name in sheets_map else sheet_names[0]
                    self._update_current_df_and_display("left", target)
            else:
                self.path_right = path
                self.right_sheets_map = sheets_map
                self._set_sheet_selector_items("right", sheet_names, default_sheet_name)
                if sheet_names:
                    target = default_sheet_name if default_sheet_name in sheets_map else sheet_names[0]
                    self._update_current_df_and_display("right", target)

            self._update_common_sheet_selector()

            self._update_file_path_label(side, path, loading=False)
            self._close_progress_dialog(dialog, progress)
        except Exception as e:
            messagebox.showerror("오류", f"데이터 적용 실패:\n{e}")
            self._close_progress_dialog(dialog, progress)

    def _show_path_tooltip(self, event, full_path):
        """전체 경로 툴팁 표시"""
        tooltip = tk.Toplevel()
        tooltip.wm_overrideredirect(True)
        tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
        label = tk.Label(tooltip, text=full_path, background="#ffffe0", 
                        relief="solid", borderwidth=1, font=("맑은 고딕", 9),
                        padx=5, pady=3)
        label.pack()
        self._current_tooltip = tooltip
    
    def _hide_tooltip(self):
        """툴팁 숨기기"""
        if hasattr(self, '_current_tooltip'):
            if self._current_tooltip.winfo_exists():
                self._current_tooltip.destroy()
            delattr(self, '_current_tooltip')
    
    def load_left(self):
        """왼쪽 파일 로드"""
        self._load_file_common("left")

    def load_right(self):
        """오른쪽 파일 로드"""
        self._load_file_common("right")


    # ============================================
    # 🔹 값 정규화 함수 (공백, 개행, 숫자 포맷 정리)
    # ============================================
    @staticmethod
    def _normalize_value(value):
        """
        셀 값을 정규화하여 비교 가능한 형태로 변환
        - None을 빈 문자열로 변환
        - 개행 문자 정리 (\r\n, \r → \n)
        - 앞뒤 공백 제거
        - 숫자 문자열은 Decimal로 변환 후 비교 (정밀도 손실 방지)
        
        Args:
            value: 정규화할 값
            
        Returns:
            정규화된 값 (숫자는 Decimal, 그 외는 문자열)
        """
        if value is None:
            return ""
        
        normalized = str(value)
        # 개행 문자 정리
        normalized = normalized.replace("\r\n", "\n").replace("\r", "\n")
        # 앞뒤 공백 제거
        normalized = normalized.strip()
        
        # 숫자 문자열인 경우 Decimal로 변환 (정밀도 손실 방지)
        try:
            # Decimal로 변환 시도 (정밀도 유지)
            decimal_value = Decimal(normalized)
            # 무한대나 NaN이 아닌 경우 Decimal 반환
            if decimal_value.is_finite():
                return decimal_value
        except (InvalidOperation, ValueError, OverflowError):
            pass
        
        # 숫자가 아니거나 변환 실패 시 문자열 반환
        return normalized

    # ============================================
    # 🔹 비교 취소 처리
    # ============================================
    def _cancel_compare(self):
        """비교 취소"""
        self._compare_cancelled = True

    # ============================================
    # 🔹 차이점 내보내기
    # ============================================
    def _export_diff_list(self, diff_list):
        """
        차이점 목록을 CSV 파일로 내보내기
        
        Args:
            diff_list: 차이점 목록 [(행, 열, Excel주소, 왼쪽값, 오른쪽값), ...]
        """
        if not diff_list:
            messagebox.showinfo("안내", "내보낼 차이점이 없습니다.")
            return
        
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="차이점 목록 저장"
        )
        
        if not path:
            return
        
        try:
            import csv
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                # 헤더
                writer.writerow(["행", "열", "Excel 주소", "왼쪽 파일 값", "오른쪽 파일 값"])
                # 데이터
                for row_idx, col_idx, excel_addr, left_val, right_val in diff_list:
                    writer.writerow([row_idx + 1, col_idx + 1, excel_addr, left_val, right_val])
            
            messagebox.showinfo("완료", f"차이점 목록이 저장되었습니다.\n파일: {path}")
        except Exception as e:
            messagebox.showerror("오류", f"파일 저장 실패:\n{e}")

    # ============================================
    # 🔹 차이점 목록 표시 다이얼로그
    # ============================================
    def _show_diff_list_dialog(self, diff_list, diff_count):
        """
        차이점 목록을 표시하는 다이얼로그
        
        Args:
            diff_list: 차이점 목록
            diff_count: 차이점 개수
        """
        dialog = tk.Toplevel(self.root)
        dialog.title(f"차이점 목록 ({diff_count}개)")
        dialog.geometry("800x600")
        dialog.transient(self.root)
        
        # 중앙 배치
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (800 // 2)
        y = (dialog.winfo_screenheight() // 2) - (600 // 2)
        dialog.geometry(f"800x600+{x}+{y}")
        
        # 상단 정보
        info_frame = tk.Frame(dialog)
        info_frame.pack(fill="x", padx=10, pady=10)
        tk.Label(info_frame, text=f"총 {diff_count}개의 차이점이 발견되었습니다.", 
                font=("맑은 고딕", 10, "bold")).pack(side=tk.LEFT)
        tk.Button(info_frame, text="CSV로 내보내기", 
                 command=lambda: self._export_diff_list(diff_list),
                 bg=self.theme_colors["button_primary"],
                 fg=self.theme_colors["button_text"]).pack(side=tk.RIGHT)
        
        # 리스트박스와 스크롤바
        list_frame = tk.Frame(dialog)
        list_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, font=("맑은 고딕", 9), yscrollcommand=scrollbar.set)
        listbox.pack(side=tk.LEFT, fill="both", expand=True)
        scrollbar.config(command=listbox.yview)
        
        # 차이점 목록 표시
        for row_idx, col_idx, excel_addr, left_val, right_val in diff_list:
            # 값이 너무 길면 잘라서 표시
            left_display = str(left_val)[:30] + "..." if len(str(left_val)) > 30 else str(left_val)
            right_display = str(right_val)[:30] + "..." if len(str(right_val)) > 30 else str(right_val)
            listbox.insert(tk.END, f"{excel_addr}: 왼쪽='{left_display}' / 오른쪽='{right_display}'")
        
        # 더블클릭 시 해당 셀로 이동
        def on_double_click(event):
            selection = listbox.curselection()
            if selection:
                idx = selection[0]
                row_idx, col_idx, _, _, _ = diff_list[idx]
                # 해당 셀로 이동
                self.left_sheet.see(row_idx, col_idx)
                self.right_sheet.see(row_idx, col_idx)
                self.left_sheet.select_cell(row_idx, col_idx)
                self.right_sheet.select_cell(row_idx, col_idx)
        
        listbox.bind("<Double-Button-1>", on_double_click)
        
        # 닫기 버튼
        tk.Button(dialog, text="닫기", command=dialog.destroy, width=15).pack(pady=10)

    # ============================================
    # 🔹 내부 비교 (차이점 강조) - 비동기 버전
    # ============================================
    def compare_files(self):
        """두 엑셀 파일을 비교하여 차이점을 표시 (비동기 처리)"""
        if self.df_left is None or self.df_right is None:
            messagebox.showwarning("안내", "두 파일을 모두 불러온 후 비교를 진행하세요.")
            return

        # 비교 취소 플래그 초기화
        self._compare_cancelled = False
        self._diff_list = []

        # 버튼 상태 변경
        self.compare_button.config(state=tk.DISABLED)
        self.cancel_button.config(state=tk.NORMAL)

        # 기존 강조 제거
        self.left_sheet.dehighlight_all()
        self.right_sheet.dehighlight_all()

        # 진행 다이얼로그 먼저 생성 (메인 스레드에서)
        dialog, progress, label = self._show_progress_dialog("파일 비교 중...")

        # 비동기 비교 시작
        thread = threading.Thread(target=self._compare_files_async, args=(dialog, progress, label), daemon=True)
        thread.start()

    def _compare_files_async(self, dialog, progress, label):
        """비교 작업 비동기 처리"""
        try:
            
            rows = min(len(self.df_left), len(self.df_right))
            cols = min(len(self.df_left.columns), len(self.df_right.columns))
            total_cells = (rows - 1) * (cols - 1)  # A열과 1행 제외
            processed = 0
            diff_count = 0
            diff_list = []

            # A열과 1행은 비교하지 않음 (헤더 영역)
            for i in range(1, rows):
                # 취소 확인
                if self._compare_cancelled:
                    self.root.after(0, lambda: self._close_progress_dialog(dialog, progress))
                    self.root.after(0, lambda: self.compare_button.config(state=tk.NORMAL))
                    self.root.after(0, lambda: self.cancel_button.config(state=tk.DISABLED))
                    self.root.after(0, lambda: messagebox.showinfo("취소", "비교가 취소되었습니다."))
                    return
                
                for j in range(1, cols):
                    # 취소 확인
                    if self._compare_cancelled:
                        self.root.after(0, lambda: self._close_progress_dialog(dialog, progress))
                        self.root.after(0, lambda: self.compare_button.config(state=tk.NORMAL))
                        self.root.after(0, lambda: self.cancel_button.config(state=tk.DISABLED))
                        self.root.after(0, lambda: messagebox.showinfo("취소", "비교가 취소되었습니다."))
                        return
                    
                    val_a = self._normalize_value(self.df_left.iat[i, j])
                    val_b = self._normalize_value(self.df_right.iat[i, j])
                    
                    # Decimal과 문자열 비교 시 타입 변환
                    if isinstance(val_a, Decimal) and isinstance(val_b, Decimal):
                        is_diff = val_a != val_b
                    elif isinstance(val_a, Decimal):
                        is_diff = str(val_a) != val_b
                    elif isinstance(val_b, Decimal):
                        is_diff = val_a != str(val_b)
                    else:
                        is_diff = val_a != val_b
                    
                    if is_diff:
                        diff_count += 1
                        # Excel 주소 계산
                        excel_addr = self._row_col_to_excel_address(i, j)
                        # 원본 값 저장 (정규화 전)
                        left_val = self.df_left.iat[i, j]
                        right_val = self.df_right.iat[i, j]
                        diff_list.append((i, j, excel_addr, left_val, right_val))
                        
                        # UI 업데이트 (메인 스레드에서) - 클로저 문제 방지를 위해 기본값 사용
                        def make_highlight(row, col):
                            def highlight():
                                self.left_sheet.highlight_cells(row=row, column=col, bg=self.theme_colors["diff_highlight"])
                                self.right_sheet.highlight_cells(row=row, column=col, bg=self.theme_colors["diff_highlight"])
                            return highlight
                        
                        self.root.after(0, make_highlight(i, j))
                    
                    processed += 1
                    # 진행률 업데이트 (100개마다)
                    if processed % 100 == 0:
                        progress_pct = (processed / total_cells * 100) if total_cells > 0 else 0
                        # 클로저 문제 방지
                        def update_progress(p, t, pct):
                            return lambda: self._update_progress_message(
                                dialog, label, f"비교 중... ({p}/{t}, {pct:.1f}%)"
                            )
                        self.root.after(0, update_progress(processed, total_cells, progress_pct))

            # 비교 완료 처리 (메인 스레드에서)
            self._diff_list = diff_list
            self.root.after(0, self._on_compare_complete, diff_count, diff_list, dialog, progress)
            
        except Exception as e:
            error_msg = f"비교 중 오류 발생:\n{e}"
            self.root.after(0, lambda: messagebox.showerror("오류", error_msg))
            if dialog:
                self.root.after(0, lambda: self._close_progress_dialog(dialog, progress))
    
    def _on_compare_complete(self, diff_count, diff_list, dialog, progress):
        """비교 완료 처리 (메인 스레드에서 실행)"""
        self._close_progress_dialog(dialog, progress)
        
        # 버튼 상태 복원
        self.compare_button.config(state=tk.NORMAL)
        self.cancel_button.config(state=tk.DISABLED)
        
        # 결과 표시
        result_msg = f"비교 완료\n값이 다른 셀 수: {diff_count}개"
        if diff_count > 0:
            result_msg += "\n\n차이점 목록을 보시겠습니까?"
            response = messagebox.askyesno("비교 완료", result_msg)
            if response:
                self._show_diff_list_dialog(diff_list, diff_count)
        else:
            messagebox.showinfo("비교 완료", result_msg)
    
    # ============================================
    # 🔹 창 설정 관련 메서드
    # ============================================
    def _setup_window_geometry(self):
        """창 크기/위치 설정"""
        window_config = self.config["window"]
        width = window_config["width"]
        height = window_config["height"]
        x = window_config["x"]
        y = window_config["y"]
        
        if x is not None and y is not None:
            self.root.geometry(f"{width}x{height}+{x}+{y}")
        else:
            # 화면 중앙에 배치
            self.root.geometry(f"{width}x{height}")
            self.root.update_idletasks()
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            self.root.geometry(f"{width}x{height}+{x}+{y}")
    
    def _save_window_geometry(self):
        """현재 창 크기/위치 저장"""
        try:
            geometry = self.root.geometry()
            # 형식: "1700x850+100+200" 또는 "1700x850"
            if '+' in geometry:
                parts = geometry.split('+')
                size_part = parts[0]
                x = int(parts[1])
                y = int(parts[2]) if len(parts) > 2 else 0
            else:
                size_part = geometry
                x = self.root.winfo_x()
                y = self.root.winfo_y()
            
            size_parts = size_part.split('x')
            if len(size_parts) == 2:
                width = int(size_parts[0])
                height = int(size_parts[1])
                self.config["window"]["width"] = width
                self.config["window"]["height"] = height
                self.config["window"]["x"] = x
                self.config["window"]["y"] = y
                ConfigManager.save_config(self.config)
        except Exception as e:
            print(f"창 크기/위치 저장 실패: {e}")
    
    def _apply_theme(self):
        """테마 적용"""
        self.root.configure(bg=self.theme_colors["bg"])
        # 모든 위젯에 테마 적용 (필요시 확장)
    
    def _toggle_theme(self):
        """테마 전환"""
        self.theme = "dark" if self.theme == "light" else "light"
        self.config["theme"] = self.theme
        self.theme_colors = self.config["themes"][self.theme]
        ConfigManager.save_config(self.config)
        
        # 테마 재적용 (간단한 방법: 창 재생성)
        messagebox.showinfo("테마 변경", f"{'다크' if self.theme == 'dark' else '라이트'}모드로 변경되었습니다.\n변경사항을 적용하려면 프로그램을 재시작하세요.")
    
    def _on_closing(self):
        """창 닫기 이벤트 핸들러"""
        self._save_window_geometry()
        self.root.destroy()


# ============================================
# 🔹 실행부
# ============================================
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparatorApp(root)
    root.mainloop()
