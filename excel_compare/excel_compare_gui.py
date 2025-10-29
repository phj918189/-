
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, filedialog
import os

def select_file(title="엑셀 파일 선택"):
    """파일 선택창 열기"""
    path = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel files", "*.xls *.xlsx")]
    )
    return path

def compare_excel_files(file1, file2, output="비교결과.xlsx"):
    # 파일 읽기
    a = pd.read_excel(file1, sheet_name=0, dtype=str).fillna("")
    b = pd.read_excel(file2, sheet_name=0, dtype=str).fillna("")

    # 크기 통일
    max_rows = max(a.shape[0], b.shape[0])
    max_cols = max(a.shape[1], b.shape[1])

    # 첫 번째 파일을 복사해 기준으로 사용
    a.to_excel(output, index=False)
    wb = load_workbook(output)
    ws = wb.active

    # 색상 정의
    green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # 셀 단위 비교
    for i in range(max_rows):
        for j in range(max_cols):
            val_a = str(a.iat[i, j]) if i < len(a.index) and j < len(a.columns) else ""
            val_b = str(b.iat[i, j]) if i < len(b.index) and j < len(b.columns) else ""

            if val_a != val_b:
                ws.cell(row=i + 2, column=j + 1).fill = green  # 헤더 아래부터 색상 변경

    wb.save(output)
    print(f"✅ 비교 완료: {output}")
    print("   - 초록색: 값이 불일치한 셀")

def main():
    root = Tk()
    root.withdraw()  # Tk 창 숨김

    print("=== 엑셀 파일 비교 프로그램 ===")
    file1 = select_file("첫 번째 파일(기준 파일)을 선택하세요")
    if not file1:
        print("❌ 첫 번째 파일이 선택되지 않았습니다.")
        return
    file2 = select_file("두 번째 파일(비교 대상 파일)을 선택하세요")
    if not file2:
        print("❌ 두 번째 파일이 선택되지 않았습니다.")
        return

    print("\n📁 기준 파일:", os.path.basename(file1))
    print("📁 비교 대상:", os.path.basename(file2))
    compare_excel_files(file1, file2)

if __name__ == "__main__":
    main()
